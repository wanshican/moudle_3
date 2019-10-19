import openpyxl

class OperateExcel:
    '操作excel数据的类'
    def __init__(self, filename):
        self.filename = filename

    def split_data(self):
        '拆分数据，并将数据写到不同的表中'
        wb = openpyxl.load_workbook(self.filename)
        sh = wb.active

        # 遍历，找到所有年份
        years = set()
        for row in range(2, sh.max_row+1):
            year = sh.cell(row=row, column=1).value[:4]
            years.add(year)
        years = sorted(list(years))

        # 遍历，找到和年份对应的数据，并填充进新的表
        for i in years:
            all_data = []
            for row in range(2, sh.max_row+1):
                data = []
                if sh.cell(row=row, column=1).value[:4] == i:
                    data.append(sh.cell(row=row, column=1).value)
                    data.append(sh.cell(row=row, column=2).value)
                    all_data.append(data)
            new_sh = wb.create_sheet(title=i)
            for row in range(1, len(all_data)+1):
                new_sh.cell(row=row, column=1).value = all_data[row-1][0]
                new_sh.cell(row=row, column=2).value = all_data[row-1][1]

        wb.save(self.filename)
        print('完成数据拆分。')

    def calculate_average(self):
        '计算平均值'
        wb = openpyxl.load_workbook(self.filename)
        for i in range(1, len(wb.sheetnames)):
            sh = wb[wb.sheetnames[i]]
            sh.cell(row=sh.max_row, column=2).value = f'=average(B1:B{sh.max_row-1})'
            sh.cell(row=sh.max_row, column=1).value = '平均分'

        wb.save(self.filename)
        print('平均值计算完成。')

def main():
    operateexcel = OperateExcel('btc.xlsx')
    operateexcel.split_data()
    operateexcel.calculate_average()


if __name__ == "__main__":
    main()
