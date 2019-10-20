# /usr/bin/env Python
# -*- coding:utf-8 -*-
# author: wanshican

import os
import sys
from PIL import Image
from openpyxl import Workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

class ImageSystem:
    '''图像处理类'''
    def __init__(self, sourse_dir=os.path.join(BASE_DIR, 'db'), target_dir=os.path.join(BASE_DIR, 'db')):
        self.sourse_dir = sourse_dir
        self.target_dir = target_dir

    def save_image_info(self):
        '''保存图片信息至excel文件中'''
        image_dir_list = []
        image_info_list = []
        # 遍历目录，找出图片文件
        for root, dirs, files in os.walk(self.sourse_dir):
            for name in files:
                filename = os.path.join(root, name)
                if os.path.splitext(filename)[1] in ['.jpg', '.png', '.bmp']:
                    image_dir_list.append(filename)
        # 获取每张图片的信息，并将信息保存至excel文件中
        if image_dir_list:
            for item in image_dir_list:
                info_list = []
                image_name = item.split('\\')[-1]
                image_size = Image.open(item).size
                info_list.append(image_name)
                info_list.append(image_size)
                image_info_list.append(info_list)
            wb = Workbook()
            sh = wb.active
            sh['a1'].value = '文件名'
            sh['b1'].value = '文件大小'
            for row in range(2, len(image_info_list)+1):
                sh.cell(row=row, column=1).value = image_info_list[row-2][0]
                sh.cell(row=row, column=2).value = str(image_info_list[row-2][1][0]) + '*' + str(image_info_list[row-2][1][1])
            wb.save(os.path.join(self.target_dir, 'image_info.xlsx'))
            print('图片信息保存成功！')
        else:
            print('当前目录没有图片文件，请先添加！')

            
    def resize(self):
        '''裁剪图像'''
        try:
            name = input('请输入要裁剪的图片名称（示例：小朋友.jpg）：')
            size = int(input('请输入要裁剪的尺寸（示例：100）：'))
            region = Image.open(os.path.join(self.sourse_dir, name)).crop((0, 0, size, size))
            region.save(os.path.join(self.target_dir, f'resize_{name}'))
            print('裁剪成功！')
        except Exception as e:
            print(e)
            print('未找到图片，请重试！')

    def rotate(self):
        '''旋转图像'''
        try:
            name = input('请输入要旋转的图片名称（示例：小朋友.jpg）:')
            angle = int(input('请输入旋转角度（示例：90）：'))
            result = Image.open(os.path.join(self.sourse_dir, name)).rotate(angle)
            result.save(os.path.join(self.target_dir, f'rotate_{name}'))
            print('旋转成功！')
        except Exception as e:
            print(e)
            print('未找到图片，请重试！')
        

def main():
    IS = ImageSystem()
    try:
        if sys.argv[1] in {'-re', '--resize'}:
            IS.resize()
        elif sys.argv[1] in {'-ro', '--rotate'}:
            IS.rotate()
    except Exception as e:
        print(e)

if __name__ == "__main__":
    main()
